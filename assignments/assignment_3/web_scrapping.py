import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

options = webdriver.ChromeOptions()
#options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

data = []

for year in range(2014, 2025):
  url = f"https://apps5.mineco.gob.pe/transparencia/Navegador/default.aspx?y={year}&ap=ActProy"
  driver.get(url)
  time.sleep(3)
  try:
    driver.switch_to.frame("frame0")
    print(f"ℹ️ Procesando año {year}")

    driver.find_element(By.ID, "ctl00_CPH1_BtnTipoGobierno").click()
    print("ℹ️ Seleccionando Gobiernos Regionales...")
    time.sleep(3)
    driver.find_element(By.ID, "tr2").click()
    time.sleep(1)

    print("ℹ️ Seleccionando Sector...")
    driver.find_element(By.ID, "ctl00_CPH1_BtnSector").click()
    time.sleep(4)
    driver.find_element(By.ID, "tr1").click()
    time.sleep(1)

    driver.find_element(By.ID, "ctl00_CPH1_BtnPliego").click()

    tr_index = 0
    while True:
      try:
        time.sleep(3)
        driver.find_element(By.ID, f"tr{tr_index}").click()
        time.sleep(2)

        # Extraer nombre de la región
        row = driver.find_element(By.ID, f"tr{tr_index}")
        tds = row.find_elements(By.TAG_NAME, "td")
        region_element = tds[1]
        region_name = region_element.text.split(":")[1].strip()
        print(f"ℹ️ Región: {region_name}")

        driver.find_element(By.ID, "ctl00_CPH1_BtnProgramaPpto").click()
        time.sleep(2)

        search_box = driver.find_element(By.ID, "ctl00_CPH1_TxtSearch")
        search_box.clear()
        search_box.send_keys("0090")
        driver.find_element(By.ID, "ctl00_CPH1_BtnSearchByCode").click()
        time.sleep(3)

        try:
          row = driver.find_element(By.ID, "tr0")
          tds = row.find_elements(By.TAG_NAME, "td")
          row_data = [td.text for td in tds]
          if len(row_data) > 0: 
            data.append([region_name, year, row_data[2], row_data[3], row_data[4], row_data[5], row_data[6], row_data[7], row_data[8], row_data[9] ])
        except Exception as e:
          print(f"⚠️ No se encontraron datos para {region_name} en {year}: {e}")

        driver.back()
        time.sleep(3)
        
        driver.switch_to.frame("frame0")
        
        tr_index += 1
      except Exception as e:
        print("No hay más regiones.")
        break 

    driver.switch_to.default_content()

  except Exception as e:
    print(f"⚠️ Error en el año {year}: {e}")

driver.quit()

excel_path = "presupuesto_0090.xlsx"
df = pd.DataFrame(data, columns=["Gobierno Regional", "Año", "PIA", "PIM", "Certificación", "Compromiso Anual", "Atención de Compromiso Anual","Devengado", "Girado", "Avance %"])
df.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

header_font = Font(color="FFFFFF", bold=True)
header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

for cell in ws[1]:
  cell.font = header_font
  cell.fill = header_fill

for column in ws.columns:
  max_length = 0
  column = list(column)
  for cell in column:
    try:
      if len(str(cell.value)) > max_length:
        max_length = len(cell.value)
    except:
      pass
  adjusted_width = (max_length + 2)
  ws.column_dimensions[column[0].column_letter].width = adjusted_width

wb.save(excel_path)

print("Excel generado correctamente! 'presupuesto_0090.xlsx'")
